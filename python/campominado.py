# import random
# from openpyxl import Workbook, load_workbook


# def config():
#     try:
#         wb = load_workbook(filename='campo.xlsx')
#         config = wb['config']
#     except:
#         wb = Workbook()
#         config = wb.create_sheet('config')
#         config.cell(column=1, row=1, value="linha")
#         config.cell(column=2, row=1, value=5)
#         config.cell(column=1, row=2, value="coluna")
#         config.cell(column=2, row=2, value=5)

#     linhas = config.cell(column=2, row=1).value
#     colunas = config.cell(column=2, row=2).value

#     wb.save('campo.xlsx')

#     return linhas, colunas


# def criar_tabuleiro(linhas, colunas):
#     return [['[ ]' for _ in range(colunas)] for _ in range(linhas)]


# def exibir_tabuleiro(tabuleiro):
#     for linha in tabuleiro:
#         print(' '.join('{:5}'.format(celula) for celula in linha))


# def adicionar_minas(tabuleiro, num_minas):
#     linhas = len(tabuleiro)
#     colunas = len(tabuleiro[0])
#     minas_adicionadas = 0

#     while minas_adicionadas < num_minas:
#         x = random.randint(0, linhas - 1)
#         y = random.randint(0, colunas - 1)

#         if tabuleiro[x][y] != '[M]':
#             tabuleiro[x][y] = '[M]'
#             minas_adicionadas += 1


# def jogar():
#     linhas, colunas = config()
#     tabuleiro = criar_tabuleiro(int(linhas), int(colunas))

#     num_minas = (int(linhas) * int(colunas)) // 10  # Número de minas pode ser ajustado
#     adicionar_minas(tabuleiro, num_minas)

#     while True:
#         exibir_tabuleiro(tabuleiro)
#         comando = input("Digite 'sair' para sair, ou 'jogar' para jogar: ").strip().lower()
        
#         if comando == 'sair':
#             break
#         elif comando == 'jogar':
#             x = int(input("Digite a linha (0-indexado): "))
#             y = int(input("Digite a coluna (0-indexado): "))

#             if tabuleiro[x][y] == '[M]':
#                 print("Game Over! Você encontrou uma mina.")
#                 break
#             else:
#                 tabuleiro[x][y] = '[O]'
#         else:
#             print("Comando inválido. Tente novamente.")


# def configurar():
#     linhas = int(input("Digite o número de linhas: "))
#     colunas = int(input("Digite o número de colunas: "))

#     wb = load_workbook(filename='campo.xlsx')
#     config = wb['config']
#     config.cell(column=2, row=1, value=linhas)
#     config.cell(column=2, row=2, value=colunas)
#     wb.save('campo.xlsx')

#     print("Configuração atualizada com sucesso.")


# def main():
#     while True:
#         escolha = input("Escolha uma opção: (jogar / configurar / sair): ").strip().lower()
#         if escolha == 'jogar':
#             jogar()
#         elif escolha == 'configurar':
#             configurar()
#         elif escolha == 'sair':
#             print("Saindo do jogo...")
#             break
#         else:
#             print("Opção inválida. Tente novamente.")

# if __name__ == "__main__":
#     main()













from openpyxl import Workbook, load_workbook

def config():
    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except FileNotFoundError:
        wb = Workbook()
        config = wb.create_sheet('config')
        config.cell(column=1, row=1, value="linha")
        config.cell(column=2, row=1, value=5)
        config.cell(column=1, row=2, value="coluna")
        config.cell(column=2, row=2, value=5)
        config.cell(column=1, row=3, value="dificuldade")
        config.cell(column=2, row=3, value=1)  # Definindo dificuldade padrão como 1

    linhas = config.cell(column=2, row=1).value
    colunas = config.cell(column=2, row=2).value
    dificuldade = config.cell(column=2, row=3).value

    wb.save('campo.xlsx')

    return linhas, colunas, dificuldade

def criarTabuleiro(quantLinha, quantColuna):
    for i in range(int(quantLinha)):
        for j in range(0, int(quantColuna)):
            print('{:5}'.format('[ ]') , end='')
        print()

def atualizar_config():
    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except FileNotFoundError:
        wb = Workbook()
        config = wb.create_sheet('config')

    print("Atualizar Configurações:")
    try:
        linha = int(input("Número de linhas: "))
        coluna = int(input("Número de colunas: "))
        dificuldade = int(input("Dificuldade (1-3): "))

        if dificuldade not in [1, 2, 3]:
            raise ValueError("Dificuldade deve ser entre 1 e 3.")

        config.cell(column=2, row=1, value=linha)
        config.cell(column=2, row=2, value=coluna)
        config.cell(column=2, row=3, value=dificuldade)
        
        wb.save('campo.xlsx')
        print("Configurações atualizadas com sucesso!")
    except ValueError as e:
        print(f"Entrada inválida: {e}")
    except Exception as e:
        print(f"Erro ao atualizar configurações: {e}")

def menu():
    while True:
        print("\nMenu:")
        print("1. Jogar")
        print("2. Configurar")
        print("3. Sair")

        escolha = input("Escolha uma opção: ")

        if escolha == '1':
            linha, coluna, dificuldade = config()
            criarTabuleiro(linha, coluna)
        elif escolha == '2':
            atualizar_config()
        elif escolha == '3':
            print("Saindo...")
            break
        else:
            print("Opção inválida. Tente novamente.")

# Iniciar o menu
menu()
