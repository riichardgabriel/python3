from openpyxl import Workbook, load_workbook
import random

def configs():
    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except:
        wb = Workbook()
        config = wb.create_sheet('config')
        config.cell(column=1, row=1, value="linha")
        config.cell(column=2, row=1, value=5)
        config.cell(column=1, row=2, value="coluna")
        config.cell(column=2, row=2, value=5)
        config.cell(column=1, row=3, value="dificuldade")
        config.cell(column=2, row=3, value=2)
        

    linhas = config.cell(column=2, row=1).value
    colunas = config.cell(column=2, row=2).value
    dificuldade = config.cell(column=2, row=3).value

    wb.save('campo.xlsx')

    config = {
        'linha' : int(linhas),
        'coluna' : int(colunas),
        'dificuldade' : int(dificuldade)
    }

    return config



def criarTabuleiro(quantLinha, quantColuna):
    for i in range(int(quantLinha)):
        for j in range(0, int(quantColuna)):
            print('{:5}'.format('[ ]') , end='')
        print()


def configCampo():
    novaQuantColuna = input('Quantas Colunas: ')
    novaQuantLinha = input('Quantas Linhas: ')
    novaDificuldade = input('1- Facil \n2- Médio \n3-Difícil \nEscolha a dificuldade: ')


    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except:
        wb = Workbook()
        config = wb.create_sheet('config')

    config.cell(column=1, row=1, value="linha")
    config.cell(column=2, row=1, value="novaQuantLinha")
    config.cell(column=1, row=2, value="coluna")
    config.cell(column=2, row=2, value="novaQuantColuna")
    config.cell(column=1, row=3, value="dificuldade")
    config.cell(column=2, row=3, value="novaDificuldade")

    wb.save('campo.xlsx')



def calculoBombas():

    config = configs()
    totalCasas = config['linha'] * config['coluna']


    if config['dificuldade'] ==1:
        quantBomba = totalCasas * 0.15
    elif config['dificuldade'] == 2:
        quantBomba = totalCasas * 0.30
    elif config['dificuldade'] == 3:
        quantBomba = totalCasas * 0.50
        
    print('total de bombas: {}'.format(int(quantBomba)))

    print(random.randint(1, 10))


    return int(quantBomba)

# config = configs()
# criarTabuleiro(config['linha'], config['coluna'])
calculoBombas()


# while True:
#     calculoBombas()
#     opcao = input('1- Jogar \n2- Configurar \n3- Sair \nEscolha uma das opções: ')
#     if opcao == '1':
#         print('Jogar')
#     elif opcao == '2':
#         print('Configurar')
#         configCampo()
#     elif opcao == '3':
#         print('Sair')
#         break
#     else:
#         print('\033[31mTu é besta?\033[0m')

# config = configs()